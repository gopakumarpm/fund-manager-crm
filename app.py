import streamlit as st
import sys
import traceback

try:
    import pandas as pd
    import openpyxl
    from openpyxl.utils import get_column_letter
    import plotly.express as px
    import plotly.graph_objects as go
    from datetime import datetime, date
    import os
    import io
    import json
    import copy
except Exception as e:
    st.error(f"Import error: {e}")
    st.code(traceback.format_exc())
    st.stop()

# --- Configuration ---
try:
    DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
except Exception:
    DATA_DIR = os.path.join(os.getcwd(), "data")

try:
    os.makedirs(DATA_DIR, exist_ok=True)
except OSError:
    # Fallback to temp directory on Streamlit Cloud if data dir not writable
    import tempfile
    DATA_DIR = os.path.join(tempfile.gettempdir(), "fund_manager_data")
    os.makedirs(DATA_DIR, exist_ok=True)

BALANCE_STORE = os.path.join(DATA_DIR, "balance_data.json")
FUND_STORE = os.path.join(DATA_DIR, "fund_data.json")

st.set_page_config(
    page_title="Fund Manager CRM",
    page_icon="💰",
    layout="wide",
    initial_sidebar_state="expanded",
)

# --- Custom CSS ---
st.markdown("""
<style>
    .main .block-container { padding-top: 1rem; max-width: 1400px; }
    div[data-testid="stMetric"] {
        background: linear-gradient(135deg, #1a1a2e 0%, #16213e 100%);
        border: 1px solid #0f3460;
        border-radius: 12px;
        padding: 16px 20px;
        box-shadow: 0 4px 15px rgba(0,0,0,0.3);
    }
    div[data-testid="stMetric"] label { color: #a8b2d1 !important; font-size: 0.85rem !important; }
    div[data-testid="stMetric"] [data-testid="stMetricValue"] { color: #64ffda !important; font-size: 1.6rem !important; font-weight: 700 !important; }
    div[data-testid="stMetric"] [data-testid="stMetricDelta"] { font-size: 0.8rem !important; }
    .stTabs [data-baseweb="tab-list"] { gap: 8px; background: #0a192f; border-radius: 10px; padding: 6px; }
    .stTabs [data-baseweb="tab"] { border-radius: 8px; color: #8892b0; font-weight: 600; padding: 10px 20px; }
    .stTabs [aria-selected="true"] { background: #112240 !important; color: #64ffda !important; }
    h1, h2, h3 { color: #ccd6f6 !important; }
    div[data-testid="stSidebar"] { background: linear-gradient(180deg, #0a192f 0%, #112240 100%); }
</style>
""", unsafe_allow_html=True)


# =====================================================================
# HELPER FUNCTIONS
# =====================================================================
def format_inr(amount):
    if amount is None:
        return "N/A"
    try:
        amount = float(amount)
    except (ValueError, TypeError):
        return str(amount)
    if amount < 0:
        return f"-{format_inr(-amount)}"
    s = f"{amount:,.2f}"
    parts = s.split(".")
    integer_part = parts[0].replace(",", "")
    decimal_part = parts[1] if len(parts) > 1 else "00"
    if len(integer_part) <= 3:
        return f"\u20b9{integer_part}.{decimal_part}"
    last3 = integer_part[-3:]
    rest = integer_part[:-3]
    groups = []
    while rest:
        groups.insert(0, rest[-2:])
        rest = rest[:-2]
    return f"\u20b9{','.join(groups)},{last3}.{decimal_part}"


def safe_float(val, default=0.0):
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def serialize_for_json(obj):
    """Convert datetime objects for JSON storage."""
    if isinstance(obj, datetime):
        return {"__datetime__": obj.isoformat()}
    elif isinstance(obj, date):
        return {"__date__": obj.isoformat()}
    return obj


def deserialize_from_json(obj):
    """Restore datetime objects from JSON."""
    if isinstance(obj, dict):
        if "__datetime__" in obj:
            return datetime.fromisoformat(obj["__datetime__"])
        if "__date__" in obj:
            return date.fromisoformat(obj["__date__"])
    return obj


def save_json(filepath, data):
    """Save data dict to JSON with datetime handling."""
    def deep_convert(obj):
        if isinstance(obj, dict):
            return {k: deep_convert(v) for k, v in obj.items()}
        elif isinstance(obj, list):
            return [deep_convert(i) for i in obj]
        elif isinstance(obj, (datetime, date)):
            return serialize_for_json(obj)
        return obj

    try:
        with open(filepath, "w") as f:
            json.dump(deep_convert(data), f, indent=2)
    except OSError:
        pass  # Silently fail on read-only filesystems


def load_json(filepath):
    """Load data dict from JSON with datetime restoration."""
    def deep_restore(obj):
        if isinstance(obj, dict):
            if "__datetime__" in obj:
                return datetime.fromisoformat(obj["__datetime__"])
            if "__date__" in obj:
                return date.fromisoformat(obj["__date__"])
            return {k: deep_restore(v) for k, v in obj.items()}
        elif isinstance(obj, list):
            return [deep_restore(i) for i in obj]
        return obj

    try:
        if os.path.exists(filepath):
            with open(filepath, "r") as f:
                data = json.load(f)
            return deep_restore(data)
    except (OSError, json.JSONDecodeError):
        pass
    return None


def parse_excel_to_dict(file_bytes, filename="file.xlsx"):
    """Parse an uploaded Excel file into our standard dict format."""
    data = {}
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        if rows:
            headers = rows[0]
            data[sheet_name] = {"headers": headers, "rows": rows[1:], "raw": rows}
    wb.close()
    return data


def export_to_excel(data_dict, sheet_configs):
    """Export data back to Excel bytes for download."""
    wb = openpyxl.Workbook()
    first = True
    for sheet_name, config in sheet_configs.items():
        if sheet_name not in data_dict:
            continue
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(sheet_name)

        sheet_data = data_dict[sheet_name]
        headers = sheet_data.get("headers", [])
        rows = sheet_data.get("rows", [])

        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=h)
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, val in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# =====================================================================
# DATA LOADING - Cloud version with upload + persistence
# =====================================================================
def get_balance_data():
    """Get balance data from session state or stored JSON."""
    if "balance_data" not in st.session_state:
        stored = load_json(BALANCE_STORE)
        if stored:
            st.session_state.balance_data = stored
        else:
            st.session_state.balance_data = None
    return st.session_state.balance_data


def get_fund_data():
    """Get fund data from session state or stored JSON."""
    if "fund_data" not in st.session_state:
        stored = load_json(FUND_STORE)
        if stored:
            st.session_state.fund_data = stored
        else:
            st.session_state.fund_data = None
    return st.session_state.fund_data


def save_balance_data(data):
    st.session_state.balance_data = data
    save_json(BALANCE_STORE, data)


def save_fund_data(data):
    st.session_state.fund_data = data
    save_json(FUND_STORE, data)


# --- Sidebar ---
with st.sidebar:
    st.markdown("## Fund Manager CRM")
    st.markdown("---")

    # File upload section
    with st.expander("Upload / Sync Excel Files", expanded=not (get_balance_data() and get_fund_data())):
        st.caption("Upload your Excel files to sync data to the cloud app.")

        bal_file = st.file_uploader("My_Balance.xlsx", type=["xlsx"], key="bal_upload")
        if bal_file:
            data = parse_excel_to_dict(bal_file.getvalue(), bal_file.name)
            save_balance_data(data)
            st.success("Balance data loaded!")
            st.rerun()

        fund_file = st.file_uploader("Fund Evaluation.xlsx", type=["xlsx"], key="fund_upload")
        if fund_file:
            data = parse_excel_to_dict(fund_file.getvalue(), fund_file.name)
            save_fund_data(data)
            st.success("Fund data loaded!")
            st.rerun()

    st.markdown("---")
    page = st.radio(
        "Navigation",
        [
            "Dashboard",
            "Balance Summary",
            "Fund Evaluation",
            "Investments",
            "Loans",
            "Credit Cards",
            "Gold & Coins",
            "Update Balances",
            "Export Data",
        ],
        index=0,
    )
    st.markdown("---")

    # Status indicators
    bal_status = "Loaded" if get_balance_data() else "Not loaded"
    fund_status = "Loaded" if get_fund_data() else "Not loaded"
    st.markdown(f"**My_Balance:** {bal_status}")
    st.markdown(f"**Fund Evaluation:** {fund_status}")
    st.markdown(f"**Last refreshed:** {datetime.now().strftime('%d %b %Y %H:%M')}")


# --- Load Data ---
balance_data = get_balance_data()
fund_data = get_fund_data()

# Check if data is loaded
if not balance_data and not fund_data and page not in ["Export Data"]:
    st.markdown("# Welcome to Fund Manager CRM")
    st.markdown("---")
    st.markdown("""
    ### Getting Started

    Upload your Excel files using the sidebar to begin:

    1. **My_Balance.xlsx** - Your balance tracking file
    2. **Fund Evaluation.xlsx** - Your fund analysis file

    Once uploaded, the data persists in the cloud. You only need to re-upload when your local Excel files change.
    """)
    st.info("Use the **Upload / Sync Excel Files** section in the sidebar to get started.")
    st.stop()


# =====================================================================
# DASHBOARD PAGE
# =====================================================================
if page == "Dashboard":
    st.markdown("# Fund Manager Dashboard")

    fund_sheet = fund_data.get("Sheet1", {}) if fund_data else {}
    fund_rows = fund_sheet.get("rows", [])

    total_fund = safe_float(fund_rows[0][1]) if fund_rows else 0
    liquid_cash = safe_float(fund_rows[0][5]) if fund_rows else 0
    non_liquid = safe_float(fund_rows[0][6]) if fund_rows else 0

    bs = balance_data.get("Balance Summary", {}) if balance_data else {}
    bs_rows = bs.get("rows", [])

    latest_total = 0
    prev_total = 0
    if bs_rows:
        for r in reversed(bs_rows):
            if r[0] is not None:
                latest_total = safe_float(r[-3]) or safe_float(r[-2])
                break
        found = 0
        for r in reversed(bs_rows):
            if r[0] is not None:
                found += 1
                if found == 2:
                    prev_total = safe_float(r[-3]) or safe_float(r[-2])
                    break

    change = latest_total - prev_total
    change_pct = (change / prev_total * 100) if prev_total else 0

    cc = balance_data.get("Credit Cards", {}) if balance_data else {}
    cc_rows = cc.get("rows", [])
    total_cc_outstanding = sum(safe_float(r[2]) for r in cc_rows if r[0])

    loan = balance_data.get("Loan", {}) if balance_data else {}
    loan_rows = loan.get("rows", [])
    total_lent = 0
    total_pending_loan = 0
    for r in loan_rows:
        if r[0] and str(r[0]).strip().lower() != "total":
            total_lent += safe_float(r[1])
            total_pending_loan += abs(safe_float(r[3]))

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Total Fund Value", format_inr(total_fund))
    with col2:
        st.metric("Liquid Cash", format_inr(liquid_cash))
    with col3:
        st.metric("Non-Liquid Fund", format_inr(non_liquid))
    with col4:
        st.metric("Balance (Latest)", format_inr(latest_total),
                  delta=f"{format_inr(change)} ({change_pct:+.1f}%)")

    st.markdown("---")

    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Credit Card Outstanding", format_inr(total_cc_outstanding))
    with col2:
        st.metric("Total Lent Out", format_inr(total_lent))
    with col3:
        st.metric("Pending Recovery", format_inr(total_pending_loan))

    st.markdown("---")

    col1, col2 = st.columns(2)
    with col1:
        st.markdown("### Liquid vs Non-Liquid Funds")
        if total_fund > 0:
            fig = go.Figure(data=[go.Pie(
                labels=["Liquid Cash", "Non-Liquid Fund"],
                values=[liquid_cash, non_liquid],
                hole=0.55,
                marker=dict(colors=["#64ffda", "#e6394a"]),
                textinfo="label+percent",
                textfont=dict(size=14, color="white"),
            )])
            fig.update_layout(
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                font=dict(color="white"), height=350,
                margin=dict(l=20, r=20, t=20, b=20), showlegend=False,
            )
            fig.add_annotation(
                text=f"<b>{format_inr(total_fund)}</b><br>Total",
                x=0.5, y=0.5, showarrow=False,
                font=dict(size=14, color="#ccd6f6"),
            )
            st.plotly_chart(fig, use_container_width=True)

    with col2:
        st.markdown("### Fund Sources Breakdown")
        source_names = []
        source_values = []
        for r in fund_rows[1:]:
            if r[0] and safe_float(r[1]) > 0:
                source_names.append(str(r[0]))
                source_values.append(safe_float(r[1]))

        if source_names:
            fig2 = go.Figure(data=[go.Bar(
                x=source_values, y=source_names, orientation="h",
                marker=dict(color=source_values, colorscale="Viridis"),
                text=[format_inr(v) for v in source_values],
                textposition="auto", textfont=dict(size=10),
            )])
            fig2.update_layout(
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                font=dict(color="#ccd6f6", size=11), height=350,
                margin=dict(l=10, r=10, t=10, b=10),
                xaxis=dict(showgrid=False, showticklabels=False),
                yaxis=dict(autorange="reversed"),
            )
            st.plotly_chart(fig2, use_container_width=True)

    # Net Worth Trend
    dates = []
    totals = []
    for r in bs_rows:
        if r[0] is not None and isinstance(r[0], datetime):
            cash_val = safe_float(r[-3]) or safe_float(r[-2])
            dates.append(r[0])
            totals.append(cash_val)

    if dates:
        st.markdown("### Net Worth Trend Over Time")
        trend_df = pd.DataFrame({"Date": dates, "Total": totals})
        fig3 = go.Figure()
        fig3.add_trace(go.Scatter(
            x=trend_df["Date"], y=trend_df["Total"],
            mode="lines+markers",
            line=dict(color="#64ffda", width=2),
            marker=dict(size=5, color="#64ffda"),
            fill="tozeroy", fillcolor="rgba(100,255,218,0.1)",
            hovertemplate="Date: %{x|%d %b %Y}<br>Total: %{y:,.0f}<extra></extra>",
        ))
        fig3.update_layout(
            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
            font=dict(color="#ccd6f6"), height=350,
            margin=dict(l=10, r=10, t=10, b=10),
            xaxis=dict(showgrid=False),
            yaxis=dict(showgrid=True, gridcolor="#233554", tickformat=","),
        )
        st.plotly_chart(fig3, use_container_width=True)


# =====================================================================
# BALANCE SUMMARY PAGE
# =====================================================================
elif page == "Balance Summary":
    st.markdown("# Balance Summary")
    bs = balance_data.get("Balance Summary", {}) if balance_data else {}
    bs_headers = bs.get("headers", [])
    bs_rows = bs.get("rows", [])

    if bs_rows:
        clean_headers = [str(h) if h else f"Col_{i}" for i, h in enumerate(bs_headers)]
        df = pd.DataFrame(bs_rows, columns=clean_headers)
        if "Date" in df.columns:
            df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
            df = df.dropna(subset=["Date"])
            df = df.sort_values("Date", ascending=False)

        st.dataframe(
            df.style.format({col: "{:,.2f}" for col in df.select_dtypes(include=["float64", "int64"]).columns}),
            use_container_width=True, height=500,
        )

        st.markdown("### Account-wise Trend")
        account_cols = [c for c in clean_headers if c not in ["Date", "Cash In hand", "Column1", f"Col_{len(clean_headers)-1}"]]
        account_cols = [c for c in account_cols if c in df.columns and df[c].dtype in ["float64", "int64"]]
        selected_accounts = st.multiselect("Select accounts to compare", account_cols, default=account_cols[:4])

        if selected_accounts and "Date" in df.columns:
            fig = go.Figure()
            colors = px.colors.qualitative.Set2
            for i, acct in enumerate(selected_accounts):
                fig.add_trace(go.Scatter(
                    x=df["Date"], y=df[acct], mode="lines+markers",
                    name=acct, line=dict(color=colors[i % len(colors)], width=2),
                    marker=dict(size=4),
                ))
            fig.update_layout(
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                font=dict(color="#ccd6f6"), height=400,
                margin=dict(l=10, r=10, t=10, b=10),
                xaxis=dict(showgrid=False),
                yaxis=dict(showgrid=True, gridcolor="#233554", tickformat=","),
                legend=dict(orientation="h", y=-0.15),
            )
            st.plotly_chart(fig, use_container_width=True)
    else:
        st.warning("No balance data loaded. Upload My_Balance.xlsx from the sidebar.")


# =====================================================================
# FUND EVALUATION PAGE
# =====================================================================
elif page == "Fund Evaluation":
    st.markdown("# Fund Evaluation & Analysis")
    fund_sheet = fund_data.get("Sheet1", {}) if fund_data else {}
    fund_rows = fund_sheet.get("rows", [])

    if fund_rows:
        total_fund = safe_float(fund_rows[0][1])
        liquid_cash = safe_float(fund_rows[0][5])
        non_liquid = safe_float(fund_rows[0][6])
        liquidity_ratio = (liquid_cash / total_fund * 100) if total_fund else 0

        col1, col2, col3, col4 = st.columns(4)
        with col1: st.metric("Total Fund", format_inr(total_fund))
        with col2: st.metric("Liquid Cash", format_inr(liquid_cash))
        with col3: st.metric("Non-Liquid", format_inr(non_liquid))
        with col4: st.metric("Liquidity Ratio", f"{liquidity_ratio:.1f}%")

        st.markdown("---")
        st.markdown("### Fund Sources")
        source_data = []
        for r in fund_rows[1:]:
            if r[0]:
                source_data.append({
                    "Source": r[0], "Total Fund": safe_float(r[1]),
                    "To Be Utilised": safe_float(r[2]), "Remaining": safe_float(r[3]),
                    "Liquid": "Yes" if r[4] == "Y" else "No",
                    "Liquid Amount": safe_float(r[5]), "Non-Liquid Amount": safe_float(r[6]),
                })
        if source_data:
            sdf = pd.DataFrame(source_data)
            st.dataframe(sdf.style.format({
                "Total Fund": "{:,.2f}", "To Be Utilised": "{:,.2f}", "Remaining": "{:,.2f}",
                "Liquid Amount": "{:,.2f}", "Non-Liquid Amount": "{:,.2f}",
            }), use_container_width=True)

        st.markdown("---")
        st.markdown("### Property & Expense Tracking")
        expense_data = []
        for r in fund_rows:
            desc = r[8] if len(r) > 8 else None
            amt = r[9] if len(r) > 9 else None
            cat = r[10] if len(r) > 10 else None
            status = r[11] if len(r) > 11 else None
            if desc and amt is not None:
                expense_data.append({"Description": desc, "Amount": safe_float(amt), "Category": cat or "", "Status": status or ""})

        if expense_data:
            edf = pd.DataFrame(expense_data)
            col1, col2 = st.columns([2, 1])
            with col1:
                st.dataframe(edf.style.format({"Amount": "{:,.2f}"}), use_container_width=True)
            with col2:
                paid = edf[edf["Status"] == "Paid"]["Amount"].sum()
                pending = edf[edf["Status"] == "Pending"]["Amount"].sum()
                st.metric("Total Paid", format_inr(paid))
                st.metric("Total Pending", format_inr(pending))

        st.markdown("---")
        st.markdown("### Property Investment Summary")
        prop_metrics = {}
        for r in fund_rows:
            desc = r[8] if len(r) > 8 else None
            amt = r[9] if len(r) > 9 else None
            if desc and amt is not None:
                prop_metrics[desc] = safe_float(amt)
        if prop_metrics:
            cols = st.columns(3)
            keys_to_show = ["Property Value - TDS", "Admin Charges", "Paid Towards Property",
                            "Balance for Property", "Loan Amount", "Own contribution towards property value",
                            "Balance for Admin", "Total Own contribution"]
            idx = 0
            for k in keys_to_show:
                if k in prop_metrics:
                    with cols[idx % 3]: st.metric(k, format_inr(prop_metrics[k]))
                    idx += 1
    else:
        st.warning("No fund data loaded. Upload Fund Evaluation.xlsx from the sidebar.")


# =====================================================================
# INVESTMENTS PAGE
# =====================================================================
elif page == "Investments":
    st.markdown("# Investment Tracker (TATA AIA ULIP)")
    inv = balance_data.get("Investments", {}) if balance_data else {}
    inv_headers = inv.get("headers", [])
    inv_rows = inv.get("rows", [])

    if inv_rows:
        clean_headers = [str(h) if h else f"Col_{i}" for i, h in enumerate(inv_headers)]
        df = pd.DataFrame(inv_rows, columns=clean_headers)
        if "Company" in df.columns:
            df = df.dropna(subset=["Company"])
        policy_col = "Policy #" if "Policy #" in df.columns else (clean_headers[3] if len(clean_headers) > 3 else None)
        policies = df[policy_col].unique().tolist() if policy_col and policy_col in df.columns else []

        st.markdown("### Policy Summary (Latest Values)")
        policy_summaries = []
        for policy_id in policies:
            pdata = df[df[policy_col] == policy_id].copy()
            pdata["Date"] = pd.to_datetime(pdata["Date"], errors="coerce")
            pdata = pdata.dropna(subset=["Date"]).sort_values("Date")
            if not pdata.empty:
                latest = pdata.iloc[-1]
                policy_summaries.append({
                    "Policy #": policy_id,
                    "Premium Date": latest.get("Premium Date", "") if "Premium Date" in pdata.columns else "",
                    "Invested Amount": safe_float(latest.get("Invested Amount")) if "Invested Amount" in pdata.columns else 0,
                    "Fund Value": safe_float(latest.get("Fund Value")) if "Fund Value" in pdata.columns else 0,
                    "Growth %": safe_float(latest.get("Growth")) * 100 if "Growth" in pdata.columns else 0,
                    "Projection": safe_float(latest.get("Projection")) if "Projection" in pdata.columns else 0,
                    "Last Tracked": latest.get("Date") if "Date" in pdata.columns else "N/A",
                })

        if policy_summaries:
            combined_invested = sum(p["Invested Amount"] for p in policy_summaries)
            combined_value = sum(p["Fund Value"] for p in policy_summaries)
            combined_growth = combined_value - combined_invested
            combined_growth_pct = (combined_growth / combined_invested * 100) if combined_invested else 0

            col1, col2, col3, col4 = st.columns(4)
            with col1: st.metric("Total Invested", format_inr(combined_invested))
            with col2: st.metric("Current Fund Value", format_inr(combined_value))
            with col3: st.metric("Total Growth", format_inr(combined_growth), delta=f"{combined_growth_pct:+.2f}%")
            with col4: st.metric("Projection", format_inr(sum(p["Projection"] for p in policy_summaries)))

            st.markdown("---")
            for ps in policy_summaries:
                with st.expander(f"Policy: {ps['Policy #']} | Premium: {ps['Premium Date']} | Value: {format_inr(ps['Fund Value'])}", expanded=True):
                    c1, c2, c3, c4 = st.columns(4)
                    with c1: st.metric("Invested", format_inr(ps["Invested Amount"]))
                    with c2: st.metric("Fund Value", format_inr(ps["Fund Value"]))
                    with c3: st.metric("Growth", f"{ps['Growth %']:+.2f}%")
                    with c4: st.metric("Last Tracked", ps["Last Tracked"].strftime("%d %b %Y") if isinstance(ps["Last Tracked"], datetime) else str(ps["Last Tracked"]))

        st.markdown("---")
        st.markdown("### Fund Value Over Time")
        view_mode = st.radio("View", ["Both Policies", "Combined Total"], horizontal=True)
        chart_df = df.copy()
        if "Date" in chart_df.columns:
            chart_df["Date"] = pd.to_datetime(chart_df["Date"], errors="coerce")
            chart_df = chart_df.dropna(subset=["Date"])
        if "Fund Value" in chart_df.columns:
            chart_df["Fund Value"] = chart_df["Fund Value"].apply(safe_float)
        if "Invested Amount" in chart_df.columns:
            chart_df["Invested Amount"] = chart_df["Invested Amount"].apply(safe_float)
        if "Date" in chart_df.columns:
            chart_df = chart_df.sort_values("Date")

        fig = go.Figure()
        if view_mode == "Both Policies":
            colors = {"U172202558": "#64ffda", "U213366421": "#ffd700"}
            for policy_id in policies:
                pdata = chart_df[chart_df[policy_col] == policy_id]
                fig.add_trace(go.Scatter(x=pdata["Date"], y=pdata["Fund Value"], mode="lines+markers",
                    name=f"{policy_id} (Fund Value)", line=dict(color=colors.get(str(policy_id), "#64ffda"), width=2), marker=dict(size=4)))
                fig.add_trace(go.Scatter(x=pdata["Date"], y=pdata["Invested Amount"], mode="lines",
                    name=f"{policy_id} (Invested)", line=dict(color=colors.get(str(policy_id), "#64ffda"), width=1, dash="dash"), opacity=0.5))
        else:
            combined = chart_df.groupby("Date").agg({"Fund Value": "sum", "Invested Amount": "sum"}).reset_index().sort_values("Date")
            fig.add_trace(go.Scatter(x=combined["Date"], y=combined["Fund Value"], mode="lines+markers",
                name="Combined Fund Value", line=dict(color="#64ffda", width=3), marker=dict(size=5), fill="tozeroy", fillcolor="rgba(100,255,218,0.1)"))
            fig.add_trace(go.Scatter(x=combined["Date"], y=combined["Invested Amount"], mode="lines",
                name="Combined Invested", line=dict(color="#ff6b6b", width=2, dash="dash")))

        fig.update_layout(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
            font=dict(color="#ccd6f6"), height=450, margin=dict(l=10, r=10, t=10, b=10),
            xaxis=dict(showgrid=False), yaxis=dict(showgrid=True, gridcolor="#233554", tickformat=","),
            legend=dict(orientation="h", y=-0.15))
        st.plotly_chart(fig, use_container_width=True)
    else:
        st.warning("No investment data loaded.")


# =====================================================================
# LOANS PAGE
# =====================================================================
elif page == "Loans":
    st.markdown("# Loan Tracker")
    loan = balance_data.get("Loan", {}) if balance_data else {}
    loan_rows = loan.get("rows", [])

    if loan_rows:
        loan_data = []
        for r in loan_rows:
            name = r[0]
            if name and str(name).strip().lower() != "total":
                lent_out = safe_float(r[1])
                paid_back = safe_float(r[2])
                outstanding = abs(safe_float(r[3]))
                repaid_pct = (paid_back / lent_out * 100) if lent_out else 0
                loan_data.append({"Name": name, "Lent Out": lent_out, "Paid Back": paid_back,
                    "Outstanding": outstanding, "Repaid %": repaid_pct,
                    "Status": "Fully Repaid" if outstanding == 0 else "Pending"})

        if loan_data:
            ldf = pd.DataFrame(loan_data)
            total_lent = ldf["Lent Out"].sum()
            total_paid_back = ldf["Paid Back"].sum()
            total_outstanding = ldf["Outstanding"].sum()
            recovery_pct = (total_paid_back / total_lent * 100) if total_lent else 0

            col1, col2, col3, col4 = st.columns(4)
            with col1: st.metric("Total Lent Out", format_inr(total_lent))
            with col2: st.metric("Total Paid Back", format_inr(total_paid_back))
            with col3: st.metric("Total Outstanding", format_inr(total_outstanding))
            with col4: st.metric("Recovery Rate", f"{recovery_pct:.1f}%")

            st.markdown("---")
            st.dataframe(ldf.style.format({"Lent Out": "{:,.2f}", "Paid Back": "{:,.2f}", "Outstanding": "{:,.2f}", "Repaid %": "{:.1f}%"}), use_container_width=True)

            pending_loans = ldf[ldf["Outstanding"] > 0].sort_values("Outstanding", ascending=True)
            if not pending_loans.empty:
                st.markdown("### Pending Recoveries")
                fig = go.Figure()
                fig.add_trace(go.Bar(name="Paid Back", y=pending_loans["Name"], x=pending_loans["Paid Back"], orientation="h", marker_color="#64ffda"))
                fig.add_trace(go.Bar(name="Outstanding", y=pending_loans["Name"], x=pending_loans["Outstanding"], orientation="h", marker_color="#ff6b6b"))
                fig.update_layout(barmode="stack", paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                    font=dict(color="#ccd6f6"), height=max(250, len(pending_loans) * 45),
                    margin=dict(l=10, r=10, t=10, b=10), xaxis=dict(showgrid=True, gridcolor="#233554", tickformat=","),
                    legend=dict(orientation="h", y=-0.15))
                st.plotly_chart(fig, use_container_width=True)

            repaid_loans = ldf[ldf["Outstanding"] == 0]
            if not repaid_loans.empty:
                st.markdown("### Fully Repaid")
                st.dataframe(repaid_loans[["Name", "Lent Out", "Paid Back"]].style.format({"Lent Out": "{:,.2f}", "Paid Back": "{:,.2f}"}), use_container_width=True)
    else:
        st.warning("No loan data loaded.")


# =====================================================================
# CREDIT CARDS PAGE
# =====================================================================
elif page == "Credit Cards":
    st.markdown("# Credit Card Tracker")
    cc = balance_data.get("Credit Cards", {}) if balance_data else {}
    cc_rows = cc.get("rows", [])

    if cc_rows:
        cc_data = []
        for r in cc_rows:
            if r[0]:
                limit = safe_float(r[1])
                outstanding = safe_float(r[2])
                available = safe_float(r[3])
                utilization = (outstanding / limit * 100) if limit else 0
                cc_data.append({"Card": r[0], "Credit Limit": limit, "Outstanding": outstanding, "Available": available, "Utilization %": utilization})

        if cc_data:
            cdf = pd.DataFrame(cc_data)
            total_limit = cdf["Credit Limit"].sum()
            total_os = cdf["Outstanding"].sum()
            total_avail = cdf["Available"].sum()
            overall_util = (total_os / total_limit * 100) if total_limit else 0

            col1, col2, col3, col4 = st.columns(4)
            with col1: st.metric("Total Limit", format_inr(total_limit))
            with col2: st.metric("Total Outstanding", format_inr(total_os))
            with col3: st.metric("Available Credit", format_inr(total_avail))
            with col4: st.metric("Overall Utilization", f"{overall_util:.1f}%")

            st.markdown("---")
            st.dataframe(cdf.style.format({"Credit Limit": "{:,.2f}", "Outstanding": "{:,.2f}", "Available": "{:,.2f}", "Utilization %": "{:.1f}%"}), use_container_width=True)

            active_cards = cdf[cdf["Outstanding"] > 0]
            if not active_cards.empty:
                st.markdown("### Card Utilization")
                fig = go.Figure()
                fig.add_trace(go.Bar(name="Outstanding", x=active_cards["Card"], y=active_cards["Outstanding"], marker_color="#ff6b6b"))
                fig.add_trace(go.Bar(name="Available", x=active_cards["Card"], y=active_cards["Available"], marker_color="#64ffda"))
                fig.update_layout(barmode="stack", paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                    font=dict(color="#ccd6f6"), height=350, margin=dict(l=10, r=10, t=10, b=10),
                    yaxis=dict(showgrid=True, gridcolor="#233554", tickformat=","), legend=dict(orientation="h", y=-0.15))
                st.plotly_chart(fig, use_container_width=True)
    else:
        st.warning("No credit card data loaded.")


# =====================================================================
# GOLD & COINS PAGE
# =====================================================================
elif page == "Gold & Coins":
    st.markdown("# Gold & Coin Tracker")
    coin = balance_data.get("Coin", {}) if balance_data else {}
    coin_headers = coin.get("headers", [])
    coin_rows = coin.get("rows", [])

    if coin_rows:
        total_weight = safe_float(coin_headers[3]) if len(coin_headers) > 3 else 0
        current_rate = safe_float(coin_headers[4]) if len(coin_headers) > 4 else 0
        total_value = safe_float(coin_headers[5]) if len(coin_headers) > 5 else 0

        col1, col2, col3 = st.columns(3)
        with col1: st.metric("Total Gold Weight", f"{total_weight:.0f} grams")
        with col2: st.metric("Current Rate/gram", format_inr(current_rate))
        with col3: st.metric("Total Gold Value", format_inr(total_value))

        st.markdown("---")
        st.markdown("### Gold Coin Inventory")
        inventory = []
        for r in coin_rows[1:]:
            if r[0] and r[0] != "Name":
                inventory.append({"Source": r[0], "Coins": int(safe_float(r[1])), "Weight per coin (g)": safe_float(r[2]), "Total Weight (g)": safe_float(r[3])})
        if inventory:
            st.dataframe(pd.DataFrame(inventory), use_container_width=True)

        st.markdown("---")
        st.markdown("### Gold Price Trend")
        price_data = []
        for r in coin_rows:
            dt = r[11] if len(r) > 11 else None
            rate = r[12] if len(r) > 12 else None
            amt = r[13] if len(r) > 13 else None
            if dt and isinstance(dt, datetime) and rate:
                price_data.append({"Date": dt, "Rate": safe_float(rate), "Value": safe_float(amt)})

        if price_data:
            pdf = pd.DataFrame(price_data).sort_values("Date")
            fig = go.Figure()
            fig.add_trace(go.Scatter(x=pdf["Date"], y=pdf["Rate"], mode="lines+markers",
                name="Gold Rate/gram", line=dict(color="#ffd700", width=3), marker=dict(size=6, color="#ffd700"),
                fill="tozeroy", fillcolor="rgba(255,215,0,0.1)"))
            fig.update_layout(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                font=dict(color="#ccd6f6"), height=350, margin=dict(l=10, r=10, t=10, b=10),
                xaxis=dict(showgrid=False), yaxis=dict(showgrid=True, gridcolor="#233554", tickformat=",", title="Rate per gram (INR)"))
            st.plotly_chart(fig, use_container_width=True)

            st.markdown("### Portfolio Value Trend")
            fig2 = go.Figure()
            fig2.add_trace(go.Scatter(x=pdf["Date"], y=pdf["Value"], mode="lines+markers",
                name="Portfolio Value", line=dict(color="#64ffda", width=2), marker=dict(size=5),
                fill="tozeroy", fillcolor="rgba(100,255,218,0.1)"))
            fig2.update_layout(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                font=dict(color="#ccd6f6"), height=300, margin=dict(l=10, r=10, t=10, b=10),
                xaxis=dict(showgrid=False), yaxis=dict(showgrid=True, gridcolor="#233554", tickformat=","))
            st.plotly_chart(fig2, use_container_width=True)
    else:
        st.warning("No gold data loaded.")


# =====================================================================
# UPDATE BALANCES PAGE
# =====================================================================
elif page == "Update Balances":
    st.markdown("# Update Your Data")
    st.markdown("Make changes here. Data auto-saves to the cloud. Use **Export Data** to download updated Excel files.")

    update_tab = st.tabs(["Add Balance Entry", "Update Fund Sources", "Update Loans", "Update Credit Cards", "Update Expenses"])

    with update_tab[0]:
        st.markdown("### Add New Balance Summary Entry")
        bs = balance_data.get("Balance Summary", {}) if balance_data else {}
        bs_headers = bs.get("headers", [])
        bs_rows = bs.get("rows", [])

        if bs_rows:
            last_entry = None
            for r in reversed(bs_rows):
                if r[0] is not None:
                    last_entry = r
                    break

            if last_entry:
                last_date = last_entry[0].strftime('%d %b %Y') if isinstance(last_entry[0], datetime) else str(last_entry[0])
                st.info(f"Last entry: {last_date} | Total: {format_inr(safe_float(last_entry[-3]) or safe_float(last_entry[-2]))}")

            entry_date = st.date_input("Entry Date", value=date.today())
            account_headers = [h for h in bs_headers if h and h not in ["Date", "Cash In hand", "Column1", None]]
            new_values = {}
            for i in range(0, len(account_headers), 4):
                cols = st.columns(4)
                for j, col in enumerate(cols):
                    idx = i + j
                    if idx < len(account_headers):
                        header = account_headers[idx]
                        default_val = safe_float(last_entry[bs_headers.index(header)]) if last_entry and header in bs_headers else 0.0
                        with col:
                            new_values[header] = st.number_input(str(header), value=default_val, step=100.0, format="%.2f", key=f"bal_{header}")

            total = sum(new_values.values())
            st.markdown(f"**Calculated Total: {format_inr(total)}**")

            if st.button("Save Balance Entry", type="primary", key="save_balance"):
                new_row = [None] * len(bs_headers)
                date_idx = bs_headers.index("Date") if "Date" in bs_headers else 0
                new_row[date_idx] = datetime(entry_date.year, entry_date.month, entry_date.day)
                for header, value in new_values.items():
                    if header in bs_headers:
                        new_row[bs_headers.index(header)] = value
                cash_idx = bs_headers.index("Cash In hand") if "Cash In hand" in bs_headers else -1
                if cash_idx >= 0: new_row[cash_idx] = total
                col1_idx = bs_headers.index("Column1") if "Column1" in bs_headers else -1
                if col1_idx >= 0: new_row[col1_idx] = total

                bs_rows.append(new_row)
                balance_data["Balance Summary"]["rows"] = bs_rows
                save_balance_data(balance_data)
                st.success("Balance entry saved!")
                st.rerun()
        else:
            st.warning("Upload My_Balance.xlsx first.")

    with update_tab[1]:
        st.markdown("### Update Fund Source Values")
        fund_sheet = fund_data.get("Sheet1", {}) if fund_data else {}
        fund_rows = fund_sheet.get("rows", [])
        fund_headers = fund_sheet.get("headers", [])

        if fund_rows:
            for i, r in enumerate(fund_rows[1:], 1):
                if r[0]:
                    with st.expander(f"{r[0]} - Current: {format_inr(safe_float(r[1]))}"):
                        c1, c2, c3 = st.columns(3)
                        with c1: new_total = st.number_input("Total Fund", value=safe_float(r[1]), step=100.0, key=f"fund_total_{i}")
                        with c2: new_utilised = st.number_input("To Be Utilised", value=safe_float(r[2]), step=100.0, key=f"fund_util_{i}")
                        with c3: is_liquid = st.selectbox("Liquid?", ["Y", "N"], index=0 if r[4] == "Y" else 1, key=f"fund_liq_{i}")

                        if st.button(f"Update {r[0]}", key=f"fund_save_{i}"):
                            fund_rows[i][1] = new_total
                            fund_rows[i][2] = new_utilised
                            fund_rows[i][3] = new_total - new_utilised
                            fund_rows[i][4] = is_liquid
                            fund_rows[i][5] = (new_total - new_utilised) if is_liquid == "Y" else 0
                            fund_rows[i][6] = 0 if is_liquid == "Y" else (new_total - new_utilised)
                            fund_rows[0][1] = sum(safe_float(r[1]) for r in fund_rows[1:] if r[0])
                            fund_rows[0][5] = sum(safe_float(r[5]) for r in fund_rows[1:] if r[0])
                            fund_rows[0][6] = sum(safe_float(r[6]) for r in fund_rows[1:] if r[0])
                            fund_rows[0][3] = fund_rows[0][1]
                            fund_data["Sheet1"]["rows"] = fund_rows
                            save_fund_data(fund_data)
                            st.success(f"{r[0]} updated!")
                            st.rerun()

    with update_tab[2]:
        st.markdown("### Update Loan Records")
        loan = balance_data.get("Loan", {}) if balance_data else {}
        loan_rows = loan.get("rows", [])

        if loan_rows:
            for i, r in enumerate(loan_rows):
                name = r[0]
                if name and str(name).strip().lower() != "total":
                    outstanding = abs(safe_float(r[3]))
                    with st.expander(f"{name} - Outstanding: {format_inr(outstanding)}"):
                        c1, c2, c3 = st.columns(3)
                        with c1: new_lent = st.number_input("Lent Out", value=safe_float(r[1]), step=100.0, key=f"loan_lent_{i}")
                        with c2: new_paid = st.number_input("Paid Back", value=safe_float(r[2]), step=100.0, key=f"loan_paid_{i}")
                        with c3: st.metric("Outstanding", format_inr(new_lent - new_paid))

                        if st.button(f"Update {name}", key=f"loan_save_{i}"):
                            loan_rows[i][1] = new_lent
                            loan_rows[i][2] = new_paid
                            loan_rows[i][3] = -(new_lent - new_paid)
                            balance_data["Loan"]["rows"] = loan_rows
                            save_balance_data(balance_data)
                            st.success(f"Loan for {name} updated!")
                            st.rerun()

    with update_tab[3]:
        st.markdown("### Update Credit Card Balances")
        cc = balance_data.get("Credit Cards", {}) if balance_data else {}
        cc_rows = cc.get("rows", [])

        if cc_rows:
            for i, r in enumerate(cc_rows):
                if r[0]:
                    with st.expander(f"{r[0]} - O/S: {format_inr(safe_float(r[2]))}"):
                        c1, c2, c3 = st.columns(3)
                        with c1: new_limit = st.number_input("Credit Limit", value=safe_float(r[1]), step=1000.0, key=f"cc_limit_{i}")
                        with c2: new_os = st.number_input("Outstanding", value=safe_float(r[2]), step=100.0, key=f"cc_os_{i}")
                        with c3: st.metric("Available", format_inr(new_limit - new_os))

                        if st.button(f"Update {r[0]}", key=f"cc_save_{i}"):
                            cc_rows[i][1] = new_limit
                            cc_rows[i][2] = new_os
                            cc_rows[i][3] = new_limit - new_os
                            balance_data["Credit Cards"]["rows"] = cc_rows
                            save_balance_data(balance_data)
                            st.success(f"{r[0]} card updated!")
                            st.rerun()

    with update_tab[4]:
        st.markdown("### Update Property Expenses")
        fund_sheet = fund_data.get("Sheet1", {}) if fund_data else {}
        fund_rows = fund_sheet.get("rows", [])

        if fund_rows:
            for i, r in enumerate(fund_rows):
                desc = r[8] if len(r) > 8 else None
                amt = r[9] if len(r) > 9 else None
                status = r[11] if len(r) > 11 else None
                if desc and amt is not None:
                    with st.expander(f"{desc} - {format_inr(safe_float(amt))} ({status or 'N/A'})"):
                        c1, c2, c3, c4 = st.columns(4)
                        with c1: new_desc = st.text_input("Description", value=str(desc), key=f"exp_desc_{i}")
                        with c2: new_amt = st.number_input("Amount", value=safe_float(amt), step=100.0, key=f"exp_amt_{i}")
                        with c3: new_cat = st.text_input("Category", value=str(r[10] or ""), key=f"exp_cat_{i}")
                        with c4:
                            status_options = ["Pending", "Paid", "Part paid"]
                            new_status = st.selectbox("Status", status_options, index=status_options.index(status) if status in status_options else 0, key=f"exp_status_{i}")

                        if st.button("Update", key=f"exp_save_{i}"):
                            fund_rows[i][8] = new_desc
                            fund_rows[i][9] = new_amt
                            fund_rows[i][10] = new_cat
                            fund_rows[i][11] = new_status
                            fund_data["Sheet1"]["rows"] = fund_rows
                            save_fund_data(fund_data)
                            st.success(f"Expense '{new_desc}' updated!")
                            st.rerun()


# =====================================================================
# EXPORT DATA PAGE
# =====================================================================
elif page == "Export Data":
    st.markdown("# Export Data")
    st.markdown("Download your updated data as Excel files. Copy them back to your iCloud folder to sync.")

    col1, col2 = st.columns(2)

    with col1:
        st.markdown("### My_Balance.xlsx")
        if balance_data:
            excel_bytes = export_to_excel(balance_data, {
                "Balance Summary": {}, "Investments": {}, "Loan": {},
                "Coin": {}, "Credit Cards": {}, "V": {},
            })
            st.download_button(
                "Download My_Balance.xlsx",
                data=excel_bytes,
                file_name="My_Balance.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary",
            )
        else:
            st.warning("No balance data loaded.")

    with col2:
        st.markdown("### Fund Evaluation.xlsx")
        if fund_data:
            excel_bytes = export_to_excel(fund_data, {
                "Sheet1": {}, "Company": {},
            })
            st.download_button(
                "Download Fund Evaluation.xlsx",
                data=excel_bytes,
                file_name="Fund Evaluation.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary",
            )
        else:
            st.warning("No fund data loaded.")

    st.markdown("---")
    st.markdown("""
    ### Workflow
    1. **Upload** your Excel files from the sidebar
    2. **Review & Update** data using the CRM tabs
    3. **Export** updated files from this page
    4. **Replace** the original files in your iCloud folder
    """)
